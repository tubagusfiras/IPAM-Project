import io
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from core.database import get_db

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from weasyprint import HTML as WeasyprintHTML

router = APIRouter(tags=["Export"])

OWNER_LABELS = {"customer":"Customer","internal":"Internal","ptp":"PTP","peering":"Peering","management":"Management","reserved":"Reserved"}
STATUS_COLORS = {"active":"FF22c55e","reserved":"FF71717a","available":"FF38e8c6","deprecated":"FFef4444"}
OWNER_COLORS  = {"customer":"FF3b82f6","internal":"FF22c55e","ptp":"FFf59e0b","peering":"FFa855f7","management":"FF0ea5e9","reserved":"FF71717a"}

def _thin_border():
    s = Side(style="thin", color="FFe2e8f0")
    return Border(left=s, right=s, top=s, bottom=s)


def _build_summary_sheet(ws, block, allocs):
    bdr = _thin_border()
    left   = Alignment(horizontal="left",   vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    used  = int(block.get("used_ips") or 0)
    total = int(block.get("total_ips") or 1)
    free  = max(0, total-used)
    pct   = round(used/total*100,1) if total else 0
    pct_color = "FFef4444" if pct>85 else "FFf59e0b" if pct>60 else "FF22c55e"

    col_widths = [22,14,20,30,16,14,14]
    for i,w in enumerate(col_widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:E1")
    ws["A1"] = str(block["prefix"])
    ws["A1"].font = Font(name="Calibri",bold=True,size=20,color="FF3b82f6")
    ws["A1"].alignment = left
    ws["A1"].fill = PatternFill("solid",start_color="FF0a0f1e")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("F1:G1")
    ws["F1"] = str(block.get("status","")).upper()
    sc = "FF22c55e" if block.get("status")=="active" else "FF71717a"
    ws["F1"].font = Font(name="Calibri",bold=True,size=11,color=sc)
    ws["F1"].fill = PatternFill("solid",start_color="FF0a0f1e")
    ws["F1"].alignment = center

    for i in range(1,8):
        ws.cell(row=2,column=i).fill=PatternFill("solid",start_color="FF0d1424")
    ws.row_dimensions[2].height=4

    for i,val in enumerate([block.get("name",""),str(block["prefix"]),
                            block.get("router",""),block.get("asn",""),
                            block.get("site_name",""),block.get("operator",""),block.get("status","")],1):
        cv = ws.cell(row=3,column=i,value=val)
        cv.font=Font(name="Courier New" if i in (2,3) else "Calibri",bold=True,size=11,
                     color="FFFFFFFF" if val else "FF334155")
        cv.fill=PatternFill("solid",start_color="FF0d1424"); cv.alignment=left
    ws.row_dimensions[2].height=13; ws.row_dimensions[3].height=22

    for i in range(1,8):
        ws.cell(row=4,column=i).fill=PatternFill("solid",start_color="FF1e293b")
    ws.row_dimensions[4].height=6

    active_c = len([a for a in allocs if a.get("status")=="active"])
    resvd_c  = len([a for a in allocs if a.get("status")=="reserved"])
    stats = [("TOTAL ALLOC",str(len(allocs)),"FFFFFFFF"),
             ("ACTIVE",str(active_c),"FF22c55e"),
             ("RESERVED",str(resvd_c),"FF71717a"),
             ("FREE IPs",f"{free:,}","FF38e8c6"),
             ("USED IPs",f"{used:,}","FF3b82f6"),
             ("TOTAL IPs",f"{total:,}","FF94a3b8"),
             ("UTILIZATION",f"{pct}%",pct_color)]
    for i,(lbl,val,col) in enumerate(stats,1):
        fill=PatternFill("solid",start_color="FF0f172a")
        cl=ws.cell(row=5,column=i,value=lbl)
        cl.font=Font(name="Calibri",bold=True,size=8,color="FF64748b")
        cl.fill=fill; cl.border=bdr; cl.alignment=left
        cv=ws.cell(row=6,column=i,value=val)
        cv.font=Font(name="Calibri",bold=True,size=18,color=col)
        cv.fill=fill; cv.border=bdr; cv.alignment=left
    ws.row_dimensions[5].height=15; ws.row_dimensions[6].height=34

    filled=max(1,round(pct/100*7)) if pct>0 else 0
    for i in range(1,8):
        ws.cell(row=7,column=i).fill=PatternFill("solid",
            start_color=pct_color if i<=filled
            else "FF1e293b")
    ws.row_dimensions[7].height=14

    for i in range(1,8):
        ws.cell(row=8,column=i).fill=PatternFill("solid",start_color="FF0a0f1e")
    ws.row_dimensions[8].height=4

    owner_counts={}
    owner_colors={"customer":"FF3b82f6","internal":"FF22c55e","ptp":"FFf59e0b",
                  "peering":"FFa855f7","management":"FF0ea5e9","reserved":"FF71717a"}
    owner_labels=OWNER_LABELS
    for a in allocs:
        o=a.get("owner_type","customer")
        owner_counts[o]=owner_counts.get(o,0)+1

    ws.cell(row=9,column=1,value="TYPE BREAKDOWN").font=Font(name="Calibri",bold=True,size=8,color="FF64748b")
    ws.cell(row=9,column=1).fill=PatternFill("solid",start_color="FF0d1424")
    ws.row_dimensions[9].height=14

    for i,(k,v) in enumerate(owner_counts.items(),1):
        oc=owner_colors.get(k,"FF94a3b8")
        fill=PatternFill("solid",start_color="FF0f172a")
        lbl=ws.cell(row=10,column=i,value=owner_labels.get(k,k))
        lbl.font=Font(name="Calibri",bold=True,size=9,color=oc)
        lbl.fill=fill; lbl.border=bdr; lbl.alignment=left
        val=ws.cell(row=11,column=i,value=v)
        val.font=Font(name="Calibri",bold=True,size=20,color=oc)
        val.fill=fill; val.border=bdr; val.alignment=left
    ws.row_dimensions[10].height=16; ws.row_dimensions[11].height=34

    ws.sheet_view.showGridLines=False


def _build_block_sheet_allocs(ws, block, allocs):
    hdr_font = Font(name="Arial", bold=True, color="FFFFFFFF", size=10)
    hdr_fill = PatternFill("solid", start_color="FF1e293b")
    bdr = _thin_border()
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Block: {block['prefix']}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FF3b82f6")
    ws["A1"].alignment = left
    ws.row_dimensions[1].height = 28

    info = [("Name",block.get("name","")),("ASN",block.get("asn","")),
            ("Router",block.get("router","")),("Operator",block.get("operator","")),
            ("Site",block.get("site_name","")),("Status",str(block.get("status","")).upper())]
    for i,(k,v) in enumerate(info):
        col = (i%3)*2+1
        row = 2+i//3
        ws.cell(row=row,column=col,value=k).font = Font(name="Arial",bold=True,size=9,color="FF94a3b8")
        ws.cell(row=row,column=col+1,value=v).font = Font(name="Arial",size=10)

    used  = int(block.get("used_ips") or 0)
    total = int(block.get("total_ips") or 1)
    pct   = round(used/total*100,1) if total else 0
    ws.merge_cells("A4:H4")
    ws["A4"] = f"Utilization: {used:,} / {total:,} IPs  ({pct}%)"
    ws["A4"].font = Font(name="Arial",bold=True,size=10,
        color="FFef4444" if pct>85 else "FFf59e0b" if pct>60 else "FF22c55e")
    ws.row_dimensions[4].height = 20

    headers   = ["#","Prefix","Usable Range","Size","Type","Customer / Description","VLAN","Status"]
    col_widths = [4,  22,      30,            8,     14,     30,                    10,    12]
    for i,(h,w) in enumerate(zip(headers,col_widths),1):
        c=ws.cell(row=6,column=i,value=h)
        c.font=hdr_font; c.fill=hdr_fill; c.border=bdr; c.alignment=Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[6].height=22

    is_v6 = ":" in str(block.get("prefix",""))

    sorted_allocs = sorted(allocs, key=lambda a: a.get("prefix",""))
    merged = []
    ai = 0
    block_prefix = str(block["prefix"])
    block_net = None
    try:
        import ipaddress as _ip
        block_net = _ip.ip_network(block_prefix, strict=False)
    except: pass
    prev_end = None
    for a in sorted_allocs:
        p = a.get("prefix","")
        try:
            if block_net and not is_v6:
                n = _ip.ip_network(p, strict=False)
                if prev_end is not None and int(prev_end) < int(n.network_address):
                    merged.append({"_free":True,"range":f"{prev_end+1} - {n.network_address-1}",
                                   "size":int(n.network_address)-int(prev_end)-1})
                prev_end = n.broadcast_address
        except: pass
        merged.append({"_free":False,**dict(a)})
    if prev_end is not None and block_net and not is_v6:
        try:
            if int(prev_end) < int(block_net.broadcast_address):
                merged.append({"_free":True,"range":f"{prev_end+1} - {block_net.broadcast_address}",
                               "size":int(block_net.broadcast_address)-int(prev_end)})
        except: pass
    if not merged:
        merged = [{"_free":False,**dict(a)} for a in sorted_allocs]

    alloc_idx=0
    for r_idx,row in enumerate(merged,7):
        ws.row_dimensions[r_idx].height=18
        if row["_free"]:
            ff = PatternFill("solid",start_color="FF0a1a0f")
            for col in range(1,9):
                c=ws.cell(row=r_idx,column=col); c.fill=ff; c.border=bdr
            ws.cell(row=r_idx,column=1,value="-").font=Font(name="Arial",size=9,color="FF334155")
            ws.cell(row=r_idx,column=2,value=row["range"]).font=Font(name="Courier New",size=9,italic=True,color="FF22c55e")
            ws.cell(row=r_idx,column=3,value="Free").font=Font(name="Arial",size=9,italic=True,color="FF22c55e")
            ws.cell(row=r_idx,column=4,value=row["size"]).font=Font(name="Arial",size=9,color="FF22c55e")
            ws.cell(row=r_idx,column=5,value="FREE").font=Font(name="Arial",size=9,color="FF22c55e")
            ws.cell(row=r_idx,column=8,value="AVAILABLE").font=Font(name="Arial",size=9,color="FF22c55e")
        else:
            alloc_idx+=1
            owner  = row.get("owner_type","customer")
            status = row.get("status","active")
            o_color = OWNER_COLORS.get(owner,"FF94a3b8")
            s_color = STATUS_COLORS.get(status,"FF94a3b8")
            rf = PatternFill("solid",start_color="FF0f172a" if alloc_idx%2==0 else "FF111827")
            try: size = 2**(32-int(row["prefix"].split("/")[1])) if not is_v6 else "-"
            except: size="-"
            for col in range(1,9):
                c=ws.cell(row=r_idx,column=col); c.fill=rf; c.border=bdr; c.alignment=left
            ws.cell(row=r_idx,column=1,value=alloc_idx).font=Font(name="Arial",size=9,color="FF94a3b8")
            ws.cell(row=r_idx,column=1).alignment=center
            ws.cell(row=r_idx,column=2,value=row["prefix"]).font=Font(name="Courier New",bold=True,size=10,color="FF3b82f6")
            if not is_v6:
                try:
                    net = _ip.ip_network(row["prefix"],strict=False)
                    ws.cell(row=r_idx,column=3,value=f"{net.network_address} - {net.broadcast_address}").font=Font(name="Courier New",size=9,color="FF94a3b8")
                except: ws.cell(row=r_idx,column=3,value="").font=Font(name="Arial",size=9)
            else:
                ws.cell(row=r_idx,column=3,value="").font=Font(name="Arial",size=9)
            ws.cell(row=r_idx,column=4,value=size).font=Font(name="Arial",size=9,color="FF94a3b8")
            ws.cell(row=r_idx,column=5,value=owner.upper()).font=Font(name="Arial",bold=True,size=9,color=o_color)
            ws.cell(row=r_idx,column=6,value=row.get("description","") or row.get("customer_name","") or "").font=Font(name="Arial",size=9)
            ws.cell(row=r_idx,column=7,value=row.get("vlan_vid","")).font=Font(name="Arial",size=9)
            ws.cell(row=r_idx,column=8,value=str(status).upper()).font=Font(name="Arial",bold=True,size=9,color=s_color)

    ws.freeze_panes="A7"
    ws.sheet_view.showGridLines=False

BLOCK_QUERY = """
    SELECT b.*, s.name AS site_name,
           CASE WHEN family(b.prefix)=4 THEN
               COALESCE(SUM(CASE WHEN a.status='active' AND a.prefix::cidr!=b.prefix
                   AND NOT EXISTS(SELECT 1 FROM allocations a2 WHERE a2.block_id=b.id
                       AND a2.id!=a.id AND a2.prefix::cidr>>a.prefix::cidr AND a2.status='active')
                   THEN (2::bigint^(32-masklen(a.prefix::cidr))) ELSE 0 END),0)::numeric
           ELSE 0 END AS used_ips,
           CASE WHEN family(b.prefix)=4 THEN (2::bigint^(32-masklen(b.prefix)))::numeric
           ELSE 0 END AS total_ips
    FROM ip_blocks b LEFT JOIN sites s ON b.site_id=s.id
    LEFT JOIN allocations a ON a.block_id=b.id
    WHERE b.id=$1::uuid GROUP BY b.id,s.name
"""

ALLOC_QUERY = """
    SELECT a.prefix::text, a.status, a.owner_type, a.description, a.notes,
           c.name AS customer_name, v.vid AS vlan_vid
    FROM allocations a
    LEFT JOIN customers c ON a.customer_id=c.id
    LEFT JOIN vlans v ON a.vlan_id=v.id
    WHERE a.block_id=$1::uuid ORDER BY a.prefix::inet
"""

ALL_BLOCKS_QUERY = """
    SELECT b.*, s.name AS site_name,
           CASE WHEN family(b.prefix)=4 THEN
               COALESCE(SUM(CASE WHEN a.status='active' AND a.prefix::cidr!=b.prefix
                   AND NOT EXISTS(SELECT 1 FROM allocations a2 WHERE a2.block_id=b.id
                       AND a2.id!=a.id AND a2.prefix::cidr>>a.prefix::cidr AND a2.status='active')
                   THEN (2::bigint^(32-masklen(a.prefix::cidr))) ELSE 0 END),0)::numeric
           ELSE 0 END AS used_ips,
           CASE WHEN family(b.prefix)=4 THEN (2::bigint^(32-masklen(b.prefix)))::numeric
           ELSE 0 END AS total_ips
    FROM ip_blocks b LEFT JOIN sites s ON b.site_id=s.id
    LEFT JOIN allocations a ON a.block_id=b.id
    GROUP BY b.id,s.name ORDER BY b.prefix::inet
"""


# ── EXCEL ENDPOINTS ────────────────────────────────────────────

@router.get("/api/v1/export/block/{block_id}", summary="Export block to Excel")
async def export_block(block_id: str, db=Depends(get_db)):
    row = await db.fetchrow(BLOCK_QUERY, block_id)
    if not row: raise HTTPException(404, "Block not found")
    allocs = await db.fetch(ALLOC_QUERY, block_id)
    allocs_list = [dict(a) for a in allocs]
    block_dict = dict(row)
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _build_summary_sheet(ws_sum, block_dict, allocs_list)
    ws_alloc = wb.create_sheet(title="Allocations")
    _build_block_sheet_allocs(ws_alloc, block_dict, allocs_list)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"IPAM_{str(row['prefix']).replace('/','_').replace('.','_')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.post("/api/v1/export/blocks", summary="Export multiple blocks to Excel")
async def export_blocks(body: dict, db=Depends(get_db)):
    block_ids = body.get("block_ids", [])
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for bid in block_ids:
        row = await db.fetchrow(BLOCK_QUERY, bid)
        if not row: continue
        allocs = await db.fetch(ALLOC_QUERY, bid)
        allocs_list = [dict(a) for a in allocs]
        block_dict = dict(row)
        ws_sum = wb.create_sheet(title=str(row["prefix"]).replace("/","_")[:28]+"_S")
        _build_summary_sheet(ws_sum, block_dict, allocs_list)
        ws_alloc = wb.create_sheet(title=str(row["prefix"]).replace("/","_")[:28]+"_A")
        _build_block_sheet_allocs(ws_alloc, block_dict, allocs_list)
    ws_sum = wb.create_sheet(title="Summary", index=0)
    hdr_font = Font(name="Arial",bold=True,color="FFFFFFFF",size=10)
    hdr_fill = PatternFill("solid",start_color="FF1e293b")
    bdr = _thin_border()
    hdrs   = ["#","Prefix","Name","ASN","Router","Site","Used IPs","Total IPs","Util %","Status"]
    widths = [4,   22,      25,    12,   20,      15,    12,        12,         10,      12]
    for i,(h,w) in enumerate(zip(hdrs,widths),1):
        c=ws_sum.cell(row=1,column=i,value=h)
        c.font=hdr_font; c.fill=hdr_fill; c.border=bdr
        c.alignment=Alignment(horizontal="center")
        ws_sum.column_dimensions[get_column_letter(i)].width=w
    ws_sum.row_dimensions[1].height=22
    all_blocks = await db.fetch(ALL_BLOCKS_QUERY)
    for i,b in enumerate(all_blocks,2):
        used=int(b["used_ips"] or 0); total=int(b["total_ips"] or 1)
        pct=round(used/total*100,1) if total else 0
        s_color="FFef4444" if pct>85 else "FFf59e0b" if pct>60 else "FF22c55e"
        rf=PatternFill("solid",start_color="FF0f172a" if i%2==0 else "FF111827")
        vals=[i-1,str(b["prefix"]),b.get("name",""),b.get("asn",""),b.get("router",""),
              b.get("site_name",""),used,total,pct,str(b.get("status","")).upper()]
        for j,v in enumerate(vals,1):
            c=ws_sum.cell(row=i,column=j,value=v)
            c.fill=rf; c.border=bdr; c.font=Font(name="Arial",size=10)
            if j==9: c.font=Font(name="Arial",bold=True,size=10,color=s_color)
        ws_sum.row_dimensions[i].height=18
    ws_sum.freeze_panes="A2"
    ws_sum.sheet_view.showGridLines=False
    buf=io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":"attachment; filename=IPAM_Export.xlsx"})


@router.get("/api/v1/export/summary", summary="Export all blocks summary to Excel")
async def export_summary(db=Depends(get_db)):
    return await export_blocks({"block_ids": []}, db)


# ── PDF HELPERS ────────────────────────────────────────────────

def _get_theme_colors(dark: bool) -> dict:
    if dark:
        return {
            "bg":         "#0f172a",
            "surface":    "#1e293b",
            "surface2":   "#334155",
            "text":       "#f1f5f9",
            "text_muted": "#94a3b8",
            "border":     "#334155",
            "th_bg":      "#0f172a",
            "th_text":    "#f1f5f9",
            "td_alt":     "#1a2744",
            "card_bg":    "#1e293b",
            "card_border":"#334155",
        }
    else:
        return {
            "bg":         "#ffffff",
            "surface":    "#f8fafc",
            "surface2":   "#f1f5f9",
            "text":       "#0f172a",
            "text_muted": "#64748b",
            "border":     "#e2e8f0",
            "th_bg":      "#1e293b",
            "th_text":    "#ffffff",
            "td_alt":     "#f8fafc",
            "card_bg":    "#ffffff",
            "card_border":"#e2e8f0",
        }


def _build_block_section(block: dict, allocs: list, theme: dict, is_first: bool = True) -> str:
    prefix    = block.get("prefix", "")
    name      = block.get("name", "") or ""
    display_name = name if name and name != prefix else ""
    site      = block.get("site_name", "") or ""
    asn       = block.get("asn", "") or ""
    router    = block.get("router", "") or ""
    status    = str(block.get("status", "")).upper()
    used      = int(block.get("used_ips", 0) or 0)
    total     = int(block.get("total_ips", 1) or 1)
    free      = total - used
    pct       = round(used / total * 100, 1) if total else 0
    bar_color = "#ef4444" if pct > 85 else "#f59e0b" if pct > 60 else "#22c55e"
    t         = theme
    page_break = "" if is_first else '<div style="page-break-before:always"></div>'
    now       = __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')

    type_colors = {
        "customer":"#3b82f6","infrastructure":"#8b5cf6",
        "ptp":"#f59e0b","peering":"#a855f7",
        "management":"#0ea5e9","reserved":"#71717a","free":"#6b7280"
    }

    rows_html = ""
    for a in allocs:
        typ  = a.get("owner_type", "") or ""
        tc   = type_colors.get(typ, "#6b7280")
        vlan = a.get('vlan_name') or (f"VID {a.get('vlan_vid')}" if a.get('vlan_vid') else '-')
        rows_html += f"""
        <tr>
          <td class="col-prefix">{a.get('prefix','')}</td>
          <td class="col-customer">{a.get('customer_name','') or '-'}</td>
          <td class="col-type"><span style="background:{tc};color:#fff;padding:2px 8px;border-radius:4px;font-size:10px">{typ}</span></td>
          <td class="col-vlan">{vlan}</td>
          <td class="col-status">{a.get('status','')}</td>
          <td class="col-desc">{a.get('description','') or '-'}</td>
        </tr>"""

    name_tag = f'<div class="name-tag">{display_name}</div>' if display_name else ''

    body = f"""{page_break}
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px">
    <div><h1>{prefix}</h1>{name_tag}</div>
    <div style="text-align:right">
      <div style="font-size:10px;color:{t['text_muted']}">{site} &nbsp;|&nbsp; {asn} &nbsp;|&nbsp; {router}</div>
      <div style="margin-top:2px"><span style="background:{bar_color};color:#fff;padding:2px 10px;border-radius:4px;font-size:11px;font-weight:700">{status}</span></div>
    </div>
  </div>
  <table class="cards-table"><tr>
    <td><div class="card-label">Total IPs</div><div class="card-value">{total:,}</div></td>
    <td><div class="card-label">Used</div><div class="card-value" style="color:#3b82f6">{used:,}</div></td>
    <td><div class="card-label">Free</div><div class="card-value" style="color:#22c55e">{free:,}</div></td>
    <td><div class="card-label">Utilization</div><div class="card-value" style="color:{bar_color}">{pct}%</div></td>
  </tr></table>
  <div class="bar-label">Utilization &mdash; {pct}%</div>
  <div class="bar-wrap"><div class="bar-fill" style="width:{pct}%;background:{bar_color}"></div></div>
  <table class="alloc-table">
    <thead><tr>
      <th class="col-prefix">Prefix</th>
      <th class="col-customer">Customer</th>
      <th class="col-type">Type</th>
      <th class="col-vlan">VLAN</th>
      <th class="col-status">Status</th>
      <th class="col-desc">End Device XC</th>
    </tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  <div class="footer">Generated by IPAM SDI &mdash; {now} &nbsp;&nbsp;|&nbsp;&nbsp; {prefix}</div>"""
    return body


def _wrap_html(body: str, theme: dict) -> str:
    t = theme
    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8"/>
<style>
  @page {{ size: A4; margin: 0; background: {t['bg']}; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  html, body {{ background: {t['bg']}; width: 100%; height: 100%; }}
  body {{ font-family: Arial, sans-serif; font-size: 12px; color: {t['text']}; padding: 15mm; }}
  h1   {{ font-size: 16px; margin: 0 0 4px 0; color: {t['text']}; font-weight: 700; }}
  .name-tag {{ font-size: 12px; font-weight: 400; color: {t['text_muted']}; }}
  .sub {{ color: {t['text_muted']}; font-size: 10px; margin-bottom: 12px; }}
  .cards-table {{ width:100%; margin-bottom:10px; }}
  .cards-table td {{ background:{t['card_bg']}; border:1px solid {t['card_border']};
                     padding:8px 12px; width:25%; text-align:center; border-radius:6px; }}
  .card-label {{ font-size:10px; color:{t['text_muted']}; margin-bottom:2px; }}
  .card-value {{ font-size:18px; font-weight:700; color:{t['text']}; }}
  .bar-label {{ font-size:10px; color:{t['text_muted']}; margin-bottom:4px; }}
  .bar-wrap {{ height:8px; background:{t['border']}; border-radius:4px; overflow:hidden; margin-bottom:14px; }}
  .bar-fill {{ height:100%; border-radius:4px; }}
  .alloc-table {{ width: 100%; border-collapse: collapse; font-size: 10px; margin-bottom: 14px; }}
  .alloc-table th {{ background: {t['th_bg']}; color: {t['th_text']}; padding: 7px 8px;
                     text-align: left; font-weight: 600; border-bottom: 2px solid {t['border']}; }}
  .alloc-table td {{ padding: 6px 8px; border-bottom: 1px solid {t['border']}; }}
  .alloc-table tr:nth-child(even) td {{ background: {t['td_alt']}; }}
  .col-prefix {{ width: 18%; }}
  .col-customer {{ width: 22%; }}
  .col-type {{ width: 12%; }}
  .col-vlan {{ width: 10%; }}
  .col-status {{ width: 10%; }}
  .col-desc {{ width: 28%; }}
  .footer {{ font-size: 9px; color: {t['text_muted']}; text-align: center; margin-top: 10px;
             border-top: 1px solid {t['border']}; padding-top: 8px; }}
</style>
</head>
<body>
{body}
</body>
</html>"""


def _build_pdf_html(block: dict, allocs: list, dark: bool = False) -> str:
    theme = _get_theme_colors(dark)
    body  = _build_block_section(block, allocs, theme, is_first=True)
    return _wrap_html(body, theme)


def _build_summary_pdf_html(all_blocks: list, dark: bool = False) -> str:
    t = _get_theme_colors(dark)
    rows_html = ""
    for i, b in enumerate(all_blocks, 1):
        used  = int(b.get("used_ips", 0) or 0)
        total = int(b.get("total_ips", 1) or 1)
        pct   = round(used / total * 100, 1) if total else 0
        color = "#ef4444" if pct > 85 else "#f59e0b" if pct > 60 else "#22c55e"
        rows_html += f"""
        <tr>
          <td>{i}</td>
          <td style="font-family:'Courier New',monospace;font-weight:700;color:#3b82f6">{b.get('prefix','')}</td>
          <td>{b.get('name','') or '-'}</td>
          <td>{b.get('asn','') or '-'}</td>
          <td>{b.get('router','') or '-'}</td>
          <td>{b.get('site_name','') or '-'}</td>
          <td style="text-align:right">{used:,}</td>
          <td style="text-align:right">{total:,}</td>
          <td style="text-align:center"><span style="background:{color};color:#fff;padding:2px 8px;border-radius:4px;font-size:10px">{pct}%</span></td>
          <td>{str(b.get('status','')).upper()}</td>
        </tr>"""
    body = f"""
  <h1>IPAM Summary</h1>
  <div class="sub">Generated by IPAM SDI &mdash; {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}</div>
  <table class="alloc-table">
    <thead><tr>
      <th>#</th><th>Prefix</th><th>Name</th><th>ASN</th><th>Router</th><th>Site</th>
      <th style="text-align:right">Used IPs</th><th style="text-align:right">Total IPs</th>
      <th style="text-align:center">Util %</th><th>Status</th>
    </tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  <div class="footer">Generated by IPAM SDI &mdash; {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}</div>"""
    return _wrap_html(body, t)


# ── PDF ENDPOINTS ──────────────────────────────────────────────

@router.get("/api/v1/export/block/{block_id}/pdf", summary="Export block to PDF")
async def export_block_pdf(block_id: str, theme: str = "dark", db=Depends(get_db)):
    row = await db.fetchrow(BLOCK_QUERY, block_id)
    if not row:
        raise HTTPException(404, "Block not found")
    allocs = await db.fetch(ALLOC_QUERY, block_id)
    allocs_list = [dict(a) for a in allocs]
    block_dict = dict(row)
    dark = theme == "dark"
    html = _build_pdf_html(block_dict, allocs_list, dark=dark)
    pdf_bytes = WeasyprintHTML(string=html).write_pdf()
    prefix_safe = str(row["prefix"]).replace("/", "_").replace(".", "_")
    fname = f"IPAM_{prefix_safe}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@router.get("/api/v1/export/summary/pdf", summary="Export summary to PDF")
async def export_summary_pdf(theme: str = "dark", db=Depends(get_db)):
    all_blocks = await db.fetch(ALL_BLOCKS_QUERY)
    all_list = [dict(b) for b in all_blocks]
    dark = theme == "dark"
    html = _build_summary_pdf_html(all_list, dark=dark)
    pdf_bytes = WeasyprintHTML(string=html).write_pdf()
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=IPAM_Summary.pdf"}
    )
