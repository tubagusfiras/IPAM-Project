from contextlib import asynccontextmanager
import asyncio, subprocess, ipaddress, time, os, json, math, io, csv
from datetime import datetime, timedelta, timezone
from typing import Optional, List

import uuid

import asyncpg
from fastapi import FastAPI, HTTPException, Query, Depends, UploadFile, File, Request, Header, BackgroundTasks
from fastapi.responses import StreamingResponse, JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from slowapi import _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from prometheus_client import Counter, Histogram, generate_latest, CONTENT_TYPE_LATEST
from loguru import logger
from core.audit import get_client_ip
import time as time_module
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from weasyprint import HTML as WeasyprintHTML
import bcrypt, jwt
import redis.asyncio as aioredis

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from weasyprint import HTML as WeasyprintHTML
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

# ── CORE MODULE IMPORTS ──────────────────────────────────────
from core.config import DATABASE_URL, REDIS_URL, JWT_SECRET, JWT_ALGORITHM, JWT_EXPIRE_HOURS, SCAN_TTL, ALLOWED_ORIGINS

# ── REDIS ────────────────────────────────────────────────────
redis_client = aioredis.from_url(REDIS_URL, decode_responses=True)
SCAN_TTL = 60 * 60 * 24  # 24 jam
from core.cache import cache_get, cache_set, cache_del

# ── SECURITY ─────────────────────────────────────────────────
security = HTTPBearer(auto_error=False)

def create_jwt_token(user_id: str, username: str, role: str) -> str:
    payload = {"sub": str(user_id), "username": username, "role": role,
               "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRE_HOURS),
               "iat": datetime.now(timezone.utc)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_jwt_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Session expired, please login again")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    if not credentials:
        raise HTTPException(401, "Not authenticated")
    return decode_jwt_token(credentials.credentials)

async def validate_api_key(key: str) -> dict | None:
    """Validate API key against stored hashes."""
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT name, key_hash, permissions FROM api_keys WHERE expires_at IS NULL OR expires_at > NOW()")
        for row in rows:
            if bcrypt.checkpw(key.encode(), row["key_hash"].encode()):
                return {"sub": f"apikey:{row['name']}", "username": f"api:{row['name']}", "role": "apikey"}
    return None

async def get_auth(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    api_key: Optional[str] = Header(None, alias="X-API-Key")
) -> dict:
    """Support both JWT (browser) and API Key (machine)."""
    if credentials:
        return decode_jwt_token(credentials.credentials)
    if api_key:
        result = await validate_api_key(api_key)
        if result:
            return result
    raise HTTPException(401, "Not authenticated")

pool: asyncpg.Pool = None

@asynccontextmanager
async def lifespan(app: FastAPI):
    global pool
    from core.database import pool as db_pool
    pool = await asyncpg.create_pool(DATABASE_URL, min_size=2, max_size=10)
    import core.database
    core.database.pool = pool
    # Start ping scheduler (di api.routes.devices)
    from api.routes.devices import _ping_scheduler
    scheduler_task = asyncio.create_task(_ping_scheduler())
    yield
    scheduler_task.cancel()
    await pool.close()

app = FastAPI(title="IPAM API", version="2.0.0", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=ALLOWED_ORIGINS, allow_methods=["*"], allow_headers=["*"], allow_credentials=True)

# Rate limiting
from core.rate_limit import limiter
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# Prometheus metrics
REQUEST_COUNT = Counter("ipam_requests_total", "Total requests", ["method", "endpoint", "status"])
REQUEST_LATENCY = Histogram("ipam_request_duration_seconds", "Request latency in seconds", ["method", "endpoint"],
                            buckets=(0.01, 0.05, 0.1, 0.25, 0.5, 1.0, 2.5, 5.0, 10.0))

# ------------------------------------------------------------------
# AUTH MIDDLEWARE
# ------------------------------------------------------------------
PUBLIC_PATHS = {"/api/v1/auth/login", "/api/v1/health/detailed", "/metrics", "/docs", "/openapi.json", "/redoc", "/api/v1/ping-trace/ping", "/api/v1/ping-trace/traceroute", "/api/v1/ping-trace/lookup", "/api/v1/ping-trace/mtr"}
PUBLIC_PREFIXES = {"/api/v1/export/block", "/api/v1/export/summary", "/api/v1/export/blocks", "/api/v1/ping/"}

@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path = request.url.path
    if path in PUBLIC_PATHS or request.method == "OPTIONS" or path == "/":
        return await call_next(request)
    if any(path.startswith(p) for p in PUBLIC_PREFIXES):
        return await call_next(request)
    if not path.startswith("/api/v1/"):
        return await call_next(request)
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        return JSONResponse(status_code=401, content={"detail": "Not authenticated"})
    try:
        request.state.user = decode_jwt_token(auth_header.replace("Bearer ", "", 1))
    except HTTPException as e:
        return JSONResponse(status_code=e.status_code, content={"detail": e.detail})
    return await call_next(request)

@app.middleware("http")
async def request_id_middleware(request: Request, call_next):
    req_id = request.headers.get("X-Request-ID") or str(uuid.uuid4())[:12]
    request.state.request_id = req_id
    response = await call_next(request)
    response.headers["X-Request-ID"] = req_id
    return response

@app.middleware("http")
async def metrics_middleware(request: Request, call_next):
    start = time_module.time()
    response = await call_next(request)
    REQUEST_COUNT.labels(request.method, request.url.path, str(response.status_code)).inc()
    REQUEST_LATENCY.labels(request.method, request.url.path).observe(time_module.time() - start)
    return response

@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    """Tambah security headers untuk mencegah serangan umum."""
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    return response

async def get_db():
    async with pool.acquire() as conn:
        yield conn


# ── ROUTE MODULE IMPORTS ─────────────────────────────────────
from models.schemas import SiteIn, CustomerIn, VlanIn, BlockIn, AllocIn
from api.routes.sites import router as sites_router
from api.routes.customers import router as customers_router
from api.routes.vlans import router as vlans_router
from api.routes.blocks import router as blocks_router
from api.routes.allocations import router as allocations_router
from api.routes.export import router as export_router
from api.routes.auth import router as auth_router
from api.routes.search import router as search_router
from api.routes.csv_import import router as csv_import_router
from api.routes.scan import router as scan_router
from api.routes.devices import router as devices_router
app.include_router(sites_router)
app.include_router(customers_router)
app.include_router(vlans_router)
app.include_router(blocks_router)
app.include_router(allocations_router)
app.include_router(export_router)
app.include_router(auth_router)
app.include_router(search_router)
app.include_router(csv_import_router)
app.include_router(scan_router)
app.include_router(devices_router)

# ------------------------------------------------------------------
# HEALTH & DASHBOARD
# ------------------------------------------------------------------
@app.get("/health", summary="Simple health check", tags=["Health"])
async def health():
    return {"status": "ok"}

@app.get("/api/v1/health/detailed", summary="Detailed health — DB + Redis status", tags=["Health"])
async def health_detailed():
    h = {"status": "ok", "timestamp": datetime.now(timezone.utc).isoformat(), "services": {}}
    try:
        await pool.fetchval("SELECT 1")
        h["services"]["database"] = {"status": "ok", "pool_size": pool.get_size(), "pool_free": pool.get_idle_size()}
    except Exception as e:
        h["services"]["database"] = {"status": "error", "error": str(e)}
        h["status"] = "degraded"
    try:
        await redis_client.ping()
        info = await redis_client.info("memory")
        h["services"]["redis"] = {"status": "ok", "used_memory_human": info.get("used_memory_human", "unknown")}
    except Exception as e:
        h["services"]["redis"] = {"status": "error", "error": str(e)}
        h["status"] = "degraded"
    h["ok_count"] = sum(1 for s in h["services"].values() if s["status"] == "ok")
    h["total_count"] = len(h["services"])
    return h

@app.get("/metrics", summary="Prometheus metrics endpoint", tags=["Monitoring"])
async def metrics():
    return Response(generate_latest(), media_type=CONTENT_TYPE_LATEST)

@app.get("/api/v1/dashboard/stats", summary="Dashboard stats — counts + utilization", tags=["Dashboard"])
async def dashboard_stats(db=Depends(get_db)):
    cached = await cache_get("dashboard:stats")
    if cached: return cached
    stats = {}
    stats["total_blocks"]      = await db.fetchval("SELECT COUNT(*) FROM ip_blocks")
    stats["total_allocations"] = await db.fetchval("SELECT COUNT(*) FROM allocations")
    stats["total_customers"]   = await db.fetchval("SELECT COUNT(*) FROM customers WHERE is_active")
    stats["total_vlans"]       = await db.fetchval("SELECT COUNT(*) FROM vlans")
    stats["total_sites"]       = await db.fetchval("SELECT COUNT(*) FROM sites")
    stats["ipv4_blocks"]       = await db.fetchval("SELECT COUNT(*) FROM ip_blocks WHERE ip_version='IPv4'")
    stats["ipv6_blocks"]       = await db.fetchval("SELECT COUNT(*) FROM ip_blocks WHERE ip_version='IPv6'")
    stats["alloc_by_status"]   = dict(await db.fetchrow("""
        SELECT
            COUNT(*) FILTER (WHERE status='active')    AS active,
            COUNT(*) FILTER (WHERE status='available') AS available,
            COUNT(*) FILTER (WHERE status='reserved')  AS reserved,
            COUNT(*) FILTER (WHERE status='deprecated') AS deprecated
        FROM allocations
    """))
    _RECENT_SQL = " ".join([
        'SELECT b.prefix::text, b.name, b.ip_version, s.name AS site_name,',
        'COUNT(a.id) AS total_allocations,',
        "COUNT(a.id) FILTER (WHERE a.status = 'active') AS active_allocations,",
        'CASE WHEN family(b.prefix) = 4 THEN',
        "COALESCE(SUM(CASE WHEN a.status = 'active' AND a.prefix::cidr != b.prefix",
        'AND NOT EXISTS (SELECT 1 FROM allocations a2 WHERE a2.block_id = b.id',
        "AND a2.id != a.id AND a2.prefix::cidr >> a.prefix::cidr AND a2.status != 'available')",
        "THEN (2::bigint ^ (32 - masklen(a.prefix::cidr))) ELSE 0 END), 0)",
        'ELSE',
        "COALESCE(SUM(CASE WHEN a.status = 'active' AND a.prefix::cidr != b.prefix",
        'AND NOT EXISTS (SELECT 1 FROM allocations a2 WHERE a2.block_id = b.id',
        "AND a2.id != a.id AND a2.prefix::cidr >> a.prefix::cidr AND a2.status != 'available')",
        "THEN (2::numeric ^ (128 - masklen(a.prefix::cidr))) ELSE 0 END), 0)",
        'END AS used_ips,',
        "CASE WHEN family(b.prefix) = 4 THEN (2::bigint ^ (32 - masklen(b.prefix)))::numeric",
        "ELSE (2::numeric ^ (128 - masklen(b.prefix)))",
        'END AS total_ips',
        'FROM ip_blocks b LEFT JOIN sites s ON b.site_id = s.id',
        'LEFT JOIN allocations a ON a.block_id = b.id',
        'GROUP BY b.id, b.prefix, b.name, b.ip_version, s.name',
        'ORDER BY b.prefix::inet LIMIT 10',
    ])
    stats["recent_blocks"] = [dict(r) for r in await db.fetch(_RECENT_SQL)]
    await cache_set("dashboard:stats", stats, ttl=30)
    return stats


# EXPORT
# ------------------------------------------------------------------

def _ip_to_int(ip):
    p = list(map(int, ip.split(".")))
    return (p[0]<<24)|(p[1]<<16)|(p[2]<<8)|p[3]

def _int_to_ip(n):
    return f"{(n>>24)&255}.{(n>>16)&255}.{(n>>8)&255}.{n&255}"

def _calc_usable(prefix):
    try:
        addr, plen = prefix.split("/")
        plen = int(plen)
        base = _ip_to_int(addr)
        size = 2**(32-plen)
        if size <= 2:
            return f"{_int_to_ip(base)} - {_int_to_ip(base+size-1)}"
        return f"{_int_to_ip(base+1)} - {_int_to_ip(base+size-2)}"
    except:
        return ""

def _calc_gaps(alloc_prefixes, block_prefix):
    try:
        addr, plen = block_prefix.split("/")
        b_start = _ip_to_int(addr)
        b_end = b_start + 2**(32-int(plen)) - 1
        sorted_p = sorted(alloc_prefixes, key=lambda p: _ip_to_int(p.split("/")[0]))
        gaps, cursor = [], b_start
        for p in sorted_p:
            a_start = _ip_to_int(p.split("/")[0])
            a_end = a_start + 2**(32-int(p.split("/")[1])) - 1
            if a_start > cursor:
                gaps.append({"range": f"{_int_to_ip(cursor)} - {_int_to_ip(a_start-1)}", "size": a_start-cursor})
            cursor = max(cursor, a_end+1)
        if cursor <= b_end:
            gaps.append({"range": f"{_int_to_ip(cursor)} - {_int_to_ip(b_end)}", "size": b_end-cursor+1})
        return gaps
    except:
        return []

# ------------------------------------------------------------------
